"""
Streaming Export Service for DealVerse OS
Handles large exports with streaming, progress tracking, and memory optimization
"""
import io
import json
import asyncio
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, AsyncGenerator, Callable
from uuid import UUID, uuid4
import logging

import pandas as pd
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.config import settings
from app.services.cache_service import cache_service

logger = logging.getLogger(__name__)


class StreamingExportService:
    """Service for streaming large exports with progress tracking"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.brand_colors = {
            'primary': '#1a2332',
            'secondary': '#0066ff', 
            'accent': '#00c896',
            'text': '#333333',
            'light_gray': '#f8f9fa'
        }
        self.chunk_size = 1000  # Process data in chunks
    
    def _generate_export_id(self) -> str:
        """Generate unique export ID for tracking"""
        return str(uuid4())
    
    def _update_progress(self, export_id: str, progress: int, status: str = "processing"):
        """Update export progress in cache"""
        if cache_service.is_available():
            progress_data = {
                "export_id": export_id,
                "progress": progress,
                "status": status,
                "timestamp": datetime.utcnow().isoformat()
            }
            cache_service.set(f"export_progress:{export_id}", progress_data, ttl=3600)
    
    def get_export_progress(self, export_id: str) -> Optional[Dict[str, Any]]:
        """Get export progress from cache"""
        if cache_service.is_available():
            return cache_service.get(f"export_progress:{export_id}")
        return None
    
    async def stream_excel_export(
        self,
        data_generator: AsyncGenerator[List[Dict], None],
        headers: List[str],
        filename: str,
        export_id: Optional[str] = None,
        progress_callback: Optional[Callable] = None
    ) -> StreamingResponse:
        """Stream Excel export for large datasets"""
        
        if not export_id:
            export_id = self._generate_export_id()
        
        async def generate_excel_stream():
            """Generate Excel file in chunks"""
            try:
                # Create temporary file
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                    temp_path = temp_file.name
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Export Data"
                
                # Add headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                self._update_progress(export_id, 10, "headers_added")
                
                row_num = 2
                total_rows = 0
                
                # Process data in chunks
                async for chunk in data_generator:
                    for row_data in chunk:
                        for col, header in enumerate(headers, 1):
                            value = row_data.get(header, "")
                            ws.cell(row=row_num, column=col, value=value)
                        row_num += 1
                        total_rows += 1
                    
                    # Update progress
                    progress = min(10 + (total_rows // self.chunk_size) * 5, 90)
                    self._update_progress(export_id, progress, "processing_data")
                    
                    if progress_callback:
                        await progress_callback(progress)
                
                # Save workbook
                wb.save(temp_path)
                self._update_progress(export_id, 95, "saving_file")
                
                # Stream file content
                with open(temp_path, 'rb') as file:
                    while True:
                        chunk = file.read(8192)  # 8KB chunks
                        if not chunk:
                            break
                        yield chunk
                
                self._update_progress(export_id, 100, "completed")
                
                # Cleanup
                Path(temp_path).unlink(missing_ok=True)
                
            except Exception as e:
                logger.error(f"Error in Excel streaming export: {e}")
                self._update_progress(export_id, 0, f"error: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
        
        return StreamingResponse(
            generate_excel_stream(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    async def stream_pdf_export(
        self,
        data_generator: AsyncGenerator[List[Dict], None],
        title: str,
        filename: str,
        export_id: Optional[str] = None,
        progress_callback: Optional[Callable] = None
    ) -> StreamingResponse:
        """Stream PDF export for large datasets"""
        
        if not export_id:
            export_id = self._generate_export_id()
        
        async def generate_pdf_stream():
            """Generate PDF file in chunks"""
            try:
                # Create temporary file
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
                    temp_path = temp_file.name
                
                # Create PDF document
                doc = SimpleDocTemplate(temp_path, pagesize=A4)
                story = []
                
                # Add title
                title_style = self.styles['Title']
                story.append(Paragraph(title, title_style))
                story.append(Spacer(1, 12))
                
                self._update_progress(export_id, 10, "title_added")
                
                total_items = 0
                
                # Process data in chunks
                async for chunk in data_generator:
                    # Convert chunk to table
                    if chunk:
                        # Create table data
                        table_data = []
                        if total_items == 0:  # Add headers for first chunk
                            headers = list(chunk[0].keys())
                            table_data.append(headers)
                        
                        for item in chunk:
                            row = [str(item.get(key, "")) for key in headers]
                            table_data.append(row)
                        
                        # Create table
                        table = Table(table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 14),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        
                        story.append(table)
                        story.append(Spacer(1, 12))
                        
                        total_items += len(chunk)
                    
                    # Update progress
                    progress = min(10 + (total_items // self.chunk_size) * 5, 90)
                    self._update_progress(export_id, progress, "processing_data")
                    
                    if progress_callback:
                        await progress_callback(progress)
                
                # Build PDF
                doc.build(story)
                self._update_progress(export_id, 95, "saving_file")
                
                # Stream file content
                with open(temp_path, 'rb') as file:
                    while True:
                        chunk = file.read(8192)  # 8KB chunks
                        if not chunk:
                            break
                        yield chunk
                
                self._update_progress(export_id, 100, "completed")
                
                # Cleanup
                Path(temp_path).unlink(missing_ok=True)
                
            except Exception as e:
                logger.error(f"Error in PDF streaming export: {e}")
                self._update_progress(export_id, 0, f"error: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
        
        return StreamingResponse(
            generate_pdf_stream(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    async def stream_csv_export(
        self,
        data_generator: AsyncGenerator[List[Dict], None],
        filename: str,
        export_id: Optional[str] = None,
        progress_callback: Optional[Callable] = None
    ) -> StreamingResponse:
        """Stream CSV export for large datasets"""
        
        if not export_id:
            export_id = self._generate_export_id()
        
        async def generate_csv_stream():
            """Generate CSV file in chunks"""
            try:
                headers_written = False
                total_rows = 0
                
                self._update_progress(export_id, 10, "starting_csv_export")
                
                async for chunk in data_generator:
                    if chunk:
                        # Convert chunk to DataFrame
                        df = pd.DataFrame(chunk)
                        
                        # Write headers only for first chunk
                        csv_data = df.to_csv(index=False, header=not headers_written)
                        headers_written = True
                        
                        yield csv_data.encode('utf-8')
                        
                        total_rows += len(chunk)
                        
                        # Update progress
                        progress = min(10 + (total_rows // self.chunk_size) * 10, 95)
                        self._update_progress(export_id, progress, "processing_csv_data")
                        
                        if progress_callback:
                            await progress_callback(progress)
                
                self._update_progress(export_id, 100, "completed")
                
            except Exception as e:
                logger.error(f"Error in CSV streaming export: {e}")
                self._update_progress(export_id, 0, f"error: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
        
        return StreamingResponse(
            generate_csv_stream(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    async def create_data_generator(
        self,
        query_function: Callable,
        total_count: int,
        chunk_size: Optional[int] = None
    ) -> AsyncGenerator[List[Dict], None]:
        """Create async generator for chunked data retrieval"""
        
        if not chunk_size:
            chunk_size = self.chunk_size
        
        offset = 0
        
        while offset < total_count:
            # Get chunk of data
            chunk = await query_function(offset=offset, limit=chunk_size)
            
            if not chunk:
                break
            
            # Convert to dict if needed
            if hasattr(chunk[0], '__dict__'):
                chunk = [item.__dict__ for item in chunk]
            
            yield chunk
            offset += len(chunk)
            
            # Small delay to prevent overwhelming the database
            await asyncio.sleep(0.01)


# Global streaming export service instance
streaming_export_service = StreamingExportService()
